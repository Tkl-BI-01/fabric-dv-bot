"""
dv_watcher.py — Fabric DV Image Embedder local watcher.

Mirrors the Procurement / Puma bot architecture:
  - Render dashboard is display + API relay only.
  - This watcher runs locally on Windows: Playwright logs into SharePoint,
    downloads images, embeds into Excel, pushes the result back to Render.
  - Login credentials and MFA codes are collected via the Render dashboard
    interaction system (same login_form / otp_form / mfa_push pattern as Puma).

Env vars (set by start_dv.ps1):
  DV_RENDER_URL      https://fabric-dv-bot.onrender.com
  DV_PUSH_API_KEY    shared push key
"""
from __future__ import annotations

import base64
import io
import json
import os
import socket
import sys
import threading
import time
import traceback
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, parse_qs, unquote

from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ── Paths ─────────────────────────────────────────────────────────────
CODE_DIR = Path(__file__).resolve().parent
LOGS_DIR = CODE_DIR / "runtime" / "logs"

# ── Config ────────────────────────────────────────────────────────────
POLL_INTERVAL       = int(os.getenv("DV_POLL_INTERVAL",  "5"))
PUSH_INTERVAL       = int(os.getenv("DV_PUSH_INTERVAL",  "3"))
_INSTANCE_LOCK_PORT = 49300   # ASOS=49291 Fab=92 Puma=93 Fabrics=94 FabInv=95 Gstar=96 Levis=97 MarketSA=98 Procurement=99 DV=300

# ── Global state ──────────────────────────────────────────────────────
shutdown_requested       = threading.Event()
_instance_lock_socket: Optional[socket.socket] = None
_state_lock              = threading.Lock()
_is_running              = False
_live: dict              = {"status": "idle", "stage": "Waiting for files...", "running": False,
                             "done": 0, "total": 0, "ok_count": 0, "error_count": 0}


# ── Logging ───────────────────────────────────────────────────────────

def log(msg: str) -> None:
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line  = f"[{stamp}]  {msg}"
    print(line, flush=True)
    try:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        with (LOGS_DIR / "dv_watcher.log").open("a", encoding="utf-8") as fh:
            fh.write(line + "\n")
    except Exception:
        pass


# ── Instance lock ─────────────────────────────────────────────────────

def _acquire_instance_lock() -> bool:
    global _instance_lock_socket
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 0)
        sock.bind(("127.0.0.1", _INSTANCE_LOCK_PORT))
        sock.listen(1)
        _instance_lock_socket = sock
        return True
    except OSError:
        return False


# ── Render HTTP helpers ───────────────────────────────────────────────

def _render_url() -> str:
    return (os.getenv("DV_RENDER_URL") or "").strip().rstrip("/")


def _api_key() -> str:
    return (os.getenv("DV_PUSH_API_KEY") or "").strip()


def _render_post(path: str, data: dict) -> Optional[dict]:
    base = _render_url()
    if not base:
        return None
    try:
        body = json.dumps(data).encode()
        req  = urllib.request.Request(
            f"{base}{path}", data=body,
            headers={"Content-Type": "application/json", "X-API-Key": _api_key()},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception as exc:
        log(f"[render] POST {path} failed: {exc}")
        return None


def _render_get(path: str) -> Optional[dict]:
    base = _render_url()
    if not base:
        return None
    try:
        req = urllib.request.Request(f"{base}{path}", headers={"X-API-Key": _api_key()})
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception as exc:
        log(f"[render] GET {path} failed: {exc}")
        return None


# ── Status / interaction helpers ──────────────────────────────────────

def _push_status(status: str, stage: str, running: bool = True, **kw) -> None:
    payload = {"status": status, "stage": stage, "running": running, **kw}
    with _state_lock:
        _live.update(payload)
    _render_post("/api/push/dv/status", payload)


def _push_interaction(status: str, type_: Optional[str], message: str,
                      display_num: str = "", error: str = "") -> None:
    _render_post("/api/push/dv/interaction", {
        "status": status, "type": type_, "message": message,
        "display_num": display_num, "error": error,
    })


def _clear_interaction() -> None:
    _push_interaction("idle", None, "")


def _wait_for_response(timeout: int = 300) -> Optional[object]:
    """Poll Render until the browser submits a response to the current interaction."""
    deadline = time.time() + timeout
    while time.time() < deadline and not shutdown_requested.is_set():
        d = _render_get("/api/dv/interaction")
        if d and d.get("status") == "answered":
            return d.get("response")
        time.sleep(2)
    return None


# ── URL helper ────────────────────────────────────────────────────────

def _direct_url(viewer_url: str) -> str:
    try:
        p      = urlparse(viewer_url)
        params = parse_qs(p.query)
        if 'id' in params:
            return 'https://tklmu.sharepoint.com' + unquote(params['id'][0])
    except Exception:
        pass
    return viewer_url


# ── Main session handler ──────────────────────────────────────────────

def run_session(session: dict) -> None:
    global _is_running
    filename    = session['filename']
    excel_bytes = base64.b64decode(session['excel_b64'])
    log(f"[session] Starting — {filename}")

    pw = browser = context = page = None
    try:

        # ── 1. Ask user for SharePoint credentials via dashboard modal ──
        _push_status("processing", "Waiting for SharePoint credentials...", running=True)
        _push_interaction(
            "waiting", "login_form",
            "Enter your SharePoint / Microsoft credentials to begin.", error="",
        )
        creds = _wait_for_response(timeout=300)
        _clear_interaction()

        if not creds or not isinstance(creds, dict):
            _push_status("error", "No credentials received. Please try again.", running=False)
            return

        email    = str(creds.get("username") or creds.get("email") or "").strip()
        password = str(creds.get("password") or "").strip()
        if not email or not password:
            _push_status("error", "Email or password missing.", running=False)
            return

        _push_status("processing", "Launching browser and logging in to SharePoint...", running=True)

        # ── 2. Playwright: log in to SharePoint ──────────────────────
        pw      = sync_playwright().start()
        browser = pw.chromium.launch(headless=True,
                                     args=['--no-sandbox', '--disable-dev-shm-usage', '--disable-gpu'])
        context = browser.new_context(viewport={'width': 1280, 'height': 800})
        page    = context.new_page()

        page.goto('https://tklmu.sharepoint.com', wait_until='domcontentloaded', timeout=30000)
        page.wait_for_url('**/microsoftonline.com/**', timeout=30000)

        _push_status("processing", "Entering email...", running=True)
        page.wait_for_selector('input[type="email"]', timeout=15000)
        page.fill('input[type="email"]', email)
        page.click('input[type="submit"]')
        page.wait_for_load_state('domcontentloaded', timeout=15000)

        _push_status("processing", "Entering password...", running=True)
        try:
            page.wait_for_selector('input[type="password"]', timeout=10000)
            page.fill('input[type="password"]', password)
            page.click('input[type="submit"]')
            page.wait_for_load_state('domcontentloaded', timeout=15000)
        except Exception:
            pass

        # ── 3. Handle MFA ─────────────────────────────────────────────
        logged_in = False
        for _attempt in range(4):
            url  = page.url
            html = page.content().lower()

            if 'sharepoint.com' in url and 'microsoftonline' not in url:
                logged_in = True
                break

            # Push / number-matching MFA
            if 'approve' in html or 'notification' in html or 'number matching' in html:
                display_num = ''
                try:
                    el = page.query_selector('[data-viewid="9"] .display-number, .displaySign')
                    if el:
                        display_num = el.inner_text().strip()
                except Exception:
                    pass
                msg = 'Open Microsoft Authenticator and approve the sign-in request.'
                if display_num:
                    msg += f'  Match this number on your phone: {display_num}'
                _push_interaction("waiting", "mfa_push", msg, display_num=display_num)
                _push_status("processing",
                             f"Waiting for MFA approval on phone...{(' Number: ' + display_num) if display_num else ''}",
                             running=True)
                # Wait for user to approve on phone and click "Login Complete" in dashboard
                _wait_for_response(timeout=90)
                _clear_interaction()
                try:
                    page.wait_for_url('**/sharepoint.com/**', timeout=10000)
                    logged_in = True
                    break
                except Exception:
                    page.wait_for_load_state('domcontentloaded', timeout=5000)
                continue

            # TOTP / SMS code entry
            if 'verification code' in html or 'otc' in html or 'code' in html:
                _push_interaction("waiting", "otp_form",
                                  "Enter the 6-digit code from your Authenticator app or SMS.")
                _push_status("processing", "Waiting for MFA code from dashboard...", running=True)
                resp = _wait_for_response(timeout=120)
                _clear_interaction()
                if resp:
                    code = ''
                    if isinstance(resp, str):
                        code = resp.strip()
                    elif isinstance(resp, dict):
                        code = str(resp.get("otp") or resp.get("code") or "").strip()
                    if code:
                        inp = page.query_selector('input[name="otc"], input[placeholder*="code" i], input[type="tel"]')
                        if inp:
                            inp.fill(code)
                        btn = page.query_selector('input[type="submit"], button[type="submit"]')
                        if btn:
                            btn.click()
                        page.wait_for_load_state('domcontentloaded', timeout=15000)
                continue

            # Unknown MFA screen — generic "Login Complete" acknowledgement
            _push_interaction("waiting", "mfa_push",
                              "Additional verification required. Complete it in your authenticator app, then click Login Complete below.")
            _push_status("processing", "Waiting for MFA completion...", running=True)
            _wait_for_response(timeout=90)
            _clear_interaction()
            page.wait_for_load_state('domcontentloaded', timeout=5000)

        if not logged_in:
            # Last chance — maybe navigation completed during MFA wait
            if 'sharepoint.com' in page.url and 'microsoftonline' not in page.url:
                logged_in = True

        if not logged_in:
            _push_status("error", "Could not log in to SharePoint. Please try again.", running=False)
            return

        # Dismiss "Stay signed in?" prompt
        try:
            btn = page.query_selector('input[type="submit"]')
            if btn:
                btn.click()
                page.wait_for_load_state('domcontentloaded', timeout=8000)
        except Exception:
            pass

        log("[session] Logged in to SharePoint successfully")
        _push_status("processing", "Logged in. Scanning Excel for image URLs...", running=True)

        # ── 4. Parse Excel for SharePoint image URLs ───────────────────
        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if 'image' not in header:
            _push_status("error", "No 'image' column found in Excel file.", running=False)
            return

        url_map = [
            {'row': r, 'url': str(ws.cell(row=r, column=header.index('image') + 1).value)}
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=header.index('image') + 1).value
            and str(ws.cell(row=r, column=header.index('image') + 1).value).startswith('http')
        ]
        total = len(url_map)
        log(f"[session] Found {total} image URLs")
        _push_status("processing", f"Downloading {total} images from SharePoint...",
                     running=True, done=0, total=total)

        # ── 5. Download images via authenticated browser context ───────
        images: dict[str, str] = {}
        errors: list[str]      = []

        for i, item in enumerate(url_map):
            row        = item['row']
            direct_url = _direct_url(item['url'])
            try:
                resp = context.request.get(direct_url, timeout=15000)
                if resp.ok:
                    images[str(row)] = base64.b64encode(resp.body()).decode()
                else:
                    errors.append(f'Row {row}: HTTP {resp.status}')
            except Exception as exc:
                errors.append(f'Row {row}: {str(exc)[:60]}')

            done = i + 1
            if done % 25 == 0 or done == total:
                _push_status("processing", f"Downloading images... {done}/{total}",
                             running=True, done=done, total=total,
                             ok_count=len(images), error_count=len(errors))

        log(f"[session] Downloaded {len(images)} images, {len(errors)} errors")

        # ── 6. Embed images into Excel ─────────────────────────────────
        _push_status("processing", "Embedding images into Excel...", running=True,
                     done=total, total=total, ok_count=len(images), error_count=len(errors))

        if 'Image Preview' in header:
            ws.delete_cols(header.index('Image Preview') + 1)
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        preview_col = ws.max_column + 1
        ws.cell(row=1, column=preview_col).value = 'Image Preview'
        ws.cell(row=1, column=preview_col).font  = Font(bold=True, name='Arial')
        ws.column_dimensions[get_column_letter(preview_col)].width = 22

        embedded = 0
        for row_str, img_b64 in images.items():
            row = int(row_str)
            try:
                pil = PILImage.open(io.BytesIO(base64.b64decode(img_b64))).convert('RGB')
                pil.thumbnail((120, 120), PILImage.LANCZOS)
                buf = io.BytesIO()
                pil.save(buf, format='PNG')
                buf.seek(0)
                ws.add_image(XLImage(buf), f'{get_column_letter(preview_col)}{row}')
                ws.row_dimensions[row].height = 90
                ws.cell(row=row, column=preview_col).alignment = Alignment(
                    horizontal='center', vertical='center')
                embedded += 1
            except Exception as exc:
                log(f"[embed] row {row}: {exc}")

        out = io.BytesIO()
        wb.save(out)
        out_b64      = base64.b64encode(out.getvalue()).decode()
        out_filename = filename.replace('.xlsx', ' - With Images.xlsx')
        log(f"[session] Embedded {embedded} images")

        # ── 7. Push completed Excel back to Render ─────────────────────
        _render_post("/api/push/dv/result", {
            "excel_b64":   out_b64,
            "filename":    out_filename,
            "ok_count":    embedded,
            "error_count": len(errors),
        })
        _push_status("done", f"Done! {embedded} images embedded into {out_filename}.",
                     running=False, done=total, total=total,
                     ok_count=embedded, error_count=len(errors))
        log(f"[session] Complete — {embedded} embedded, {len(errors)} errors")

    except Exception:
        log(f"[session] ERROR:\n{traceback.format_exc()}")
        _push_status("error", "An unexpected error occurred. Check the watcher log.", running=False)
        _clear_interaction()
    finally:
        try:
            if browser:
                browser.close()
            if pw:
                pw.stop()
        except Exception:
            pass
        with _state_lock:
            _is_running = False


# ── Poll loop ─────────────────────────────────────────────────────────

def poll_loop() -> None:
    global _is_running
    log("[poll] started")
    while not shutdown_requested.is_set():
        try:
            with _state_lock:
                busy = _is_running
            if not busy:
                data = _render_get("/api/dv/pending-sessions")
                if data and data.get("sessions"):
                    session = data["sessions"][0]
                    rid     = session["request_id"]
                    _render_post("/api/dv/session-ack", {"request_id": rid})
                    with _state_lock:
                        _is_running = True
                    log(f"[poll] picked up session {rid} — {session['filename']}")
                    t = threading.Thread(target=run_session, args=(session,), daemon=True)
                    t.start()
        except Exception as exc:
            log(f"[poll] error: {exc}")
        shutdown_requested.wait(POLL_INTERVAL)


def status_push_loop() -> None:
    log("[status-push] started")
    while not shutdown_requested.is_set():
        try:
            with _state_lock:
                snap = dict(_live)
            if snap.get('running'):
                _render_post("/api/push/dv/status", snap)
        except Exception:
            pass
        shutdown_requested.wait(PUSH_INTERVAL)


# ── Entry point ───────────────────────────────────────────────────────

def main() -> None:
    if not _acquire_instance_lock():
        print("ERROR: Another DV watcher is already running on this machine.", flush=True)
        sys.exit(1)

    log("=" * 60)
    log("  DV Watcher  —  Fabric DV Image Embedder")
    log(f"  Code dir   : {CODE_DIR}")
    log(f"  Render URL : {_render_url() or '(not set — set DV_RENDER_URL)'}")
    log(f"  API key    : {'SET' if _api_key() else 'MISSING — pushes will be rejected'}")
    log("=" * 60)

    threads = [
        threading.Thread(target=poll_loop,        daemon=True, name="poll"),
        threading.Thread(target=status_push_loop, daemon=True, name="push"),
    ]
    for t in threads:
        t.start()

    _push_status("idle", "Watcher running — waiting for files...", running=False)

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log("Shutting down DV watcher...")
        shutdown_requested.set()


if __name__ == "__main__":
    main()
