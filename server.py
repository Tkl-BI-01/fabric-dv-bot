from flask import Flask, request, send_file, jsonify
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from urllib.parse import urlparse, parse_qs, unquote
import io, base64, threading, os

app = Flask(__name__, static_folder='.', static_url_path='')

PORT = int(os.environ.get('PORT', 5000))

state = {
    'status': 'idle',       # idle | logging_in | waiting_mfa | logged_in | downloading | done | error
    'mfa_type': None,       # push | totp | sms | None
    'display_num': '',      # number shown on screen for number-matching MFA
    'message': '',
    'done': 0,
    'total': 0,
    'images': {},
    'errors': [],
}
_pw = None
_browser = None
_context = None
_page = None
_lock = threading.Lock()


def parse_direct_url(viewer_url):
    try:
        p = urlparse(viewer_url)
        params = parse_qs(p.query)
        if 'id' in params:
            return 'https://tklmu.sharepoint.com' + unquote(params['id'][0])
    except Exception:
        pass
    return viewer_url

def set_state(**kwargs):
    with _lock:
        state.update(kwargs)


# ── Login ─────────────────────────────────────────────────────────────
def do_login(email, password):
    global _pw, _browser, _context, _page

    set_state(status='logging_in', message='Opening browser and navigating to SharePoint...')
    try:
        _pw = sync_playwright().start()
        launch_args = ['--no-sandbox', '--disable-dev-shm-usage', '--disable-gpu']
        _browser = _pw.chromium.launch(headless=True, args=launch_args)
        _context = _browser.new_context(viewport={'width': 1280, 'height': 800})
        _page = _context.new_page()

        _page.goto('https://tklmu.sharepoint.com', wait_until='domcontentloaded', timeout=30000)
        _page.wait_for_url('**/microsoftonline.com/**', timeout=30000)

        set_state(message='Entering email...')
        _page.wait_for_selector('input[type="email"]', timeout=15000)
        _page.fill('input[type="email"]', email)
        _page.click('input[type="submit"]')
        _page.wait_for_load_state('domcontentloaded', timeout=15000)

        set_state(message='Entering password...')
        try:
            _page.wait_for_selector('input[type="password"]', timeout=10000)
            _page.fill('input[type="password"]', password)
            _page.click('input[type="submit"]')
            _page.wait_for_load_state('domcontentloaded', timeout=15000)
        except Exception:
            pass

        _detect_post_login()

    except Exception as e:
        set_state(status='error', message=f'Login error: {str(e)}')


def _detect_post_login():
    global _page
    try:
        url = _page.url
        html = _page.content().lower()

        if 'sharepoint.com' in url and 'microsoftonline' not in url:
            _handle_stay_signed_in()
            return

        # Push / number-matching MFA
        if 'approve' in html or 'notification' in html or 'number matching' in html:
            display_num = ''
            try:
                num_el = _page.query_selector('[data-viewid="9"] .display-number, .displaySign')
                if num_el:
                    display_num = num_el.inner_text().strip()
            except Exception:
                pass
            set_state(
                status='waiting_mfa',
                mfa_type='push',
                display_num=display_num,
                message='Open Microsoft Authenticator and approve the sign-in request.'
                        + (f' Match this number: {display_num}' if display_num else ''),
            )
            try:
                _page.wait_for_url('**/sharepoint.com/**', timeout=90000)
                _handle_stay_signed_in()
            except Exception:
                set_state(status='error', message='Push notification timed out. Please try again.')
            return

        # TOTP / SMS code entry
        if 'verification code' in html or 'otc' in html or 'code' in html:
            mfa_type = 'sms' if ('text' in html or 'sms' in html or 'phone' in html) else 'totp'
            set_state(
                status='waiting_mfa',
                mfa_type=mfa_type,
                display_num='',
                message='Enter the 6-digit code from your Authenticator app or SMS.',
            )
            return

        # Unknown MFA screen
        set_state(
            status='waiting_mfa',
            mfa_type='totp',
            display_num='',
            message='Additional verification required. Enter your authentication code.',
        )

    except Exception as e:
        set_state(status='error', message=f'Error detecting login state: {e}')


def _handle_stay_signed_in():
    global _page
    try:
        btn = _page.query_selector('input[type="submit"]')
        if btn:
            btn.click()
            _page.wait_for_load_state('domcontentloaded', timeout=10000)
    except Exception:
        pass
    set_state(status='logged_in', message='Logged in successfully! Ready to download images.')


# ── Download ──────────────────────────────────────────────────────────
def do_download(url_map):
    global _context
    set_state(status='downloading', done=0, total=len(url_map), images={}, errors=[])
    images = {}
    errors = []
    try:
        for i, item in enumerate(url_map):
            row = item['row']
            direct_url = parse_direct_url(item['url'])
            try:
                resp = _context.request.get(direct_url, timeout=15000)
                if resp.ok:
                    images[str(row)] = base64.b64encode(resp.body()).decode()
                else:
                    errors.append(f'Row {row}: HTTP {resp.status}')
            except Exception as e:
                errors.append(f'Row {row}: {str(e)[:60]}')
            set_state(done=i + 1, images=images, errors=errors)
        set_state(status='done', message=f'Done! {len(images)} images downloaded.')
    except Exception as e:
        set_state(status='error', message=f'Download error: {e}')


# ── Routes ────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return app.send_static_file('image_tool.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '')
    if not email or not password:
        return jsonify({'error': 'Email and password required'}), 400
    t = threading.Thread(target=do_login, args=(email, password), daemon=True)
    t.start()
    return jsonify({'ok': True})

@app.route('/submit-mfa', methods=['POST'])
def submit_mfa():
    global _page
    code = request.json.get('code', '').strip()
    try:
        inp = _page.query_selector(
            'input[name="otc"], input[data-viewid], input[placeholder*="code" i], input[type="tel"]'
        )
        if inp:
            inp.fill(code)
        btn = _page.query_selector('input[type="submit"], button[type="submit"]')
        if btn:
            btn.click()
        _page.wait_for_load_state('domcontentloaded', timeout=15000)
        _detect_post_login()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/start-download', methods=['POST'])
def start_download():
    url_map = request.json.get('urls', [])
    t = threading.Thread(target=do_download, args=(url_map,), daemon=True)
    t.start()
    return jsonify({'ok': True})

@app.route('/status')
def get_status():
    with _lock:
        return jsonify({
            'status': state['status'],
            'mfa_type': state['mfa_type'],
            'display_num': state['display_num'],
            'message': state['message'],
            'done': state['done'],
            'total': state['total'],
            'ok_count': len(state['images']),
            'error_count': len(state['errors']),
        })

@app.route('/result')
def get_result():
    with _lock:
        return jsonify({'images': state['images'], 'errors': state['errors'][:10]})

@app.route('/embed', methods=['POST'])
def embed():
    data = request.json
    excel_bytes = base64.b64decode(data['excel'])
    images = data['images']
    filename = data.get('filename', 'Output.xlsx')

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    if 'Image Preview' in header:
        ws.delete_cols(header.index('Image Preview') + 1)
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    preview_col = ws.max_column + 1
    ws.cell(row=1, column=preview_col).value = 'Image Preview'
    ws.cell(row=1, column=preview_col).font = Font(bold=True, name='Arial')
    ws.column_dimensions[get_column_letter(preview_col)].width = 22

    embedded = 0
    for row_str, img_b64 in images.items():
        row = int(row_str)
        try:
            pil = PILImage.open(io.BytesIO(base64.b64decode(img_b64))).convert('RGB')
            pil.thumbnail((120, 120), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format='PNG')
            buf.seek(0)
            ws.add_image(XLImage(buf), f'{get_column_letter(preview_col)}{row}')
            ws.row_dimensions[row].height = 90
            ws.cell(row=row, column=preview_col).alignment = Alignment(horizontal='center', vertical='center')
            embedded += 1
        except Exception as e:
            print(f'Embed row {row}: {e}')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        download_name=filename.replace('.xlsx', ' - With Images.xlsx'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
    )


if __name__ == '__main__':
    print(f'\n{"="*50}\n  Open in Chrome: http://localhost:{PORT}\n{"="*50}\n')
    app.run(host='0.0.0.0', port=PORT, debug=False, threaded=True)
